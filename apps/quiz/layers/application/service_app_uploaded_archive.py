from datetime import datetime
from django.conf import settings
from apps.srea.layers.application.catalog_item_app_service import CatalogoItemAppService
from apps.quiz.models import Quiz, Question, Answer
from apps.srea.models import FileUploaded
import os
from django.shortcuts import render, redirect, get_object_or_404
from apps.srea.models import *

class ArchivoCargadoAppService(object):

    @staticmethod
    def pre_procesar_datos():
        FileUploaded.objects.all().delete()

    @staticmethod
    def procesar_datos(archivo_cargado): #unidad_id
        file_path = settings.MEDIA_ROOT + archivo_cargado.file.name 

        from openpyxl import load_workbook
        ws = load_workbook(file_path, data_only=True, read_only=True)

        # Obtener la hoja de cálculo deseada QUIZ
        quiz = procesar_quiz(ws)
        procesar_questions_and_anwers(ws, quiz)
        #unidad=get_object_or_404(Unidad, id=unidad_id)
        #print(unidad_id,'.....>TOBY')
        #quiz_unidad(quiz,unidad_id)
    




def procesar_quiz(object):
    try:
        for row in object._sheets[0].iter_rows(min_col=1, min_row=2,  max_col=8, max_row=2, values_only=True):
            print("quiz -> ", row)
            return Quiz.objects.create(
                    name=row[1],
                    description=row[2],
                    attempts=row[3],
                    start_date=row[4],
                    end_date=row[5],
                    time_limit=row[6],
                )
    except Exception as e:
        print('ERROR %s' % e)


def procesar_questions_and_anwers(object, quiz):
    CHOICE_TYPE = {"Multiple": "M", "Unica": "S", "Texto": "T"}
    try:
        for row in object._sheets[1].iter_rows(min_col=1, min_row=2,  max_col=4, values_only=True):
            if row[1] != None:
                question = Question(
                    statement=row[1],
                    type_question= CHOICE_TYPE[row[2]] if row[2] in CHOICE_TYPE else CHOICE_TYPE["Unica"],
                    quiz=quiz
                    )
                question.save()
                procesar_anwers(object, row[0], question)
    except Exception as e:
        print('ERROR %s' % e)



def procesar_anwers(object, order, question):
    style_list = CatalogoItemAppService.get_catalogo_item_lista('ESTILOS_APRENDIZAJE')
    try:
        for row in object._sheets[2].iter_rows(min_col=1, min_row=2,  max_col=4, values_only=True):
            if row[2] == order:     
                anwers = Answer(
                    text=row[0],
                    learning_style=[item for item in style_list if item.name == row[1]][0],
                    question=question
                )
                anwers.save()
    except Exception as e:
        print('ERROR %s' % e)


def quiz_unidad(object, unidad_id):
    unidad=get_object_or_404(Unidad, id=unidad_id)
    unidad.quizzes.add(object)
    unidad.save()